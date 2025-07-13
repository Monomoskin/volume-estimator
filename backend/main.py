from flask import Flask, request, jsonify, render_template
import pandas as pd
from PIL import Image
import io
import os
import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from prediction import volume_prediction
from flask_cors import CORS
app = Flask(__name__)
CORS(app)
EXCEL_FILE = "cells_data.xlsx"
import os

UPLOAD_FOLDER = "uploads"
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
# Función auxiliar: crea el archivo si no existe
EXCEL_FILE = "datos.xlsx"

def create_excel_if_not_exists():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cells"
        # Opcional: añade encabezados de columna si quieres
        ws.append(["cell_id", "name", "creation_date"])
        wb.save(EXCEL_FILE)
    else:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        if "Cells" not in wb.sheetnames:
            ws = wb.create_sheet("Cells")
            ws.append(["cell_id", "name", "creation_date"])
            wb.save(EXCEL_FILE)

@app.route("/")
def index():
    return render_template("index.html")

# API 1: Crear nueva célula
@app.route("/add_cell", methods=["POST"])
def add_cell():
    data = request.get_json()
    cell_id = data.get("id")
    name = data.get("name")
    created_at = data.get("created_at")

    if not all([cell_id, name, created_at]):
        return jsonify({"error": "Faltan datos"}), 400

    create_excel_if_not_exists()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Cells"]

    # Verificar si ya existe esa celula
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == cell_id:
            return jsonify({"error": "ID ya registrado"}), 400

    ws.append([cell_id, name, created_at])
    wb.save(EXCEL_FILE)
    return jsonify({"message": "Célula agregada correctamente"})

# API 2: Agregar nueva foto + volumen a una célula existente

@app.route("/add_photo", methods=["POST"])
def add_photo():
    if "image" not in request.files or "zoom" not in request.form or "cell_id" not in request.form or "timestamp" not in request.form:
        return jsonify({"error": "Faltan datos"}), 400

    image_file = request.files["image"]
    zoom_level = float(request.form["zoom"])
    cell_id = request.form["cell_id"]
    timestamp = request.form["timestamp"]

    # Guardar la imagen en disco con nombre único
    filename = f"{cell_id}_{timestamp.replace(':', '-')}.jpg"  # ':' puede ser problemático en nombres de archivos
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    image_file.save(filepath)  # guardamos archivo físico

    # Abrir la imagen para hacer la predicción
    image = Image.open(filepath).convert("RGB")
    estimated_volume = volume_prediction(image, zoom_level)

    create_excel_if_not_exists()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Cells"]

    # Buscar fila correspondiente a la célula
    row_index = None
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].value == cell_id:
            row_index = i
            break

    if not row_index:
        return jsonify({"error": "Célula no encontrada"}), 404

    # Buscar siguiente columna vacía después de las 3 primeras
    col = 4
    while ws.cell(row=row_index, column=col).value is not None:
        col += 3

    # Guardar nombre de la imagen, fecha y volumen en Excel
    ws.cell(row=row_index, column=col, value=filename)  # guardamos el nombre archivo
    ws.cell(row=row_index, column=col + 1, value=timestamp)
    ws.cell(row=row_index, column=col + 2, value=estimated_volume)

    wb.save(EXCEL_FILE)

    return jsonify({"volume": estimated_volume, "filename": filename})

# Ruta para servir las imágenes guardadas
from flask import send_from_directory

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

# API 3: Obtener los datos de una célula para graficar
@app.route("/get_cell/<cell_id>", methods=["GET"])
def get_cell(cell_id):
    create_excel_if_not_exists()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Cells"]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == cell_id:
            name = row[1]
            created_at = row[2]
            photos = []

            # Extraer pares de datos: nombre_foto, fecha, volumen
            i = 3
            while i + 2 < len(row) and row[i] is not None:
                photos.append({
                    "label": row[i],
                    "date": row[i + 1],
                    "volume": row[i + 2],
                })
                i += 3

            return jsonify({
                "id": cell_id,
                "name": name,
                "created_at": created_at,
                "photos": photos
            })

    return jsonify({"error": "Célula no encontrada"}), 404

# API 4:Borrar un registro Completo de una celula
@app.route("/delete-cell/<cell_id>", methods=["DELETE"])
def delete_cell(cell_id):
    create_excel_if_not_exists()  # Asegura que el archivo exista
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Cells"]

    row_to_delete = None
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if str(row[0].value) == str(cell_id):
            row_to_delete = i
            break

    if not row_to_delete:
        return jsonify({"error": "Célula no encontrada"}), 404

    ws.delete_rows(row_to_delete, 1)  # Elimina la fila completa
    wb.save(EXCEL_FILE)

    return jsonify({"message": "Célula eliminada correctamente"}), 200

#Get all cells
@app.route("/get_cells", methods=["GET"])
def get_cells():
    create_excel_if_not_exists()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Cells"]

    cells = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        cells.append({
            "cellId": row[0],
            "name": row[1],
            "creation_date": row[2],
        })

    return jsonify(cells)

# API de predicción original
@app.route("/predict", methods=["POST"])
def predict():
    if "image" not in request.files or "zoom" not in request.form:
        return jsonify({"error": "Missing image or zoom level"}), 400
    print("received")

    image_file = request.files["image"]
    zoom_level = float(request.form["zoom"])

    image = Image.open(io.BytesIO(image_file.read())).convert("RGB")
    estimated_volume = volume_prediction(image, zoom_level)

    return jsonify({"volume": estimated_volume})

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5001, debug=True)
